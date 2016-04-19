# -*- coding: utf-8 -*-

import openpyxl
import json
from openpyxl.styles import Font
from uuid import UUID

from django.db.models import QuerySet, Model


class XLSXSimple(object):
    """ Класс генерации xlsx из массива данных
    """
    sheet_name = None
    fields_header = []
    fields_data = [[]]

    def __init__(self, sheet_name=None, fields_header=None, fields_data=None):
        """ Инициализация класса массивами данных
        :param sheet_name: имя листа
        :param fields_header: список заголовков
        :param fields_data: матрица данных
        :return:
        """
        self.sheet_name = sheet_name or ""
        self.fields_header = fields_header or []
        self.fields_data = fields_data or [[]]

    def _get_sheet_name(self):
        return self.sheet_name

    def _get_header(self):
        return self.fields_header

    def _get_data(self):
        return self.fields_data

    @staticmethod
    def _get_field_value(field):
        if isinstance(field, UUID):
            return str(field)

        return field

    def get_work_book(self):
        """ Генерация xlsx файла
        :return: экземпляр openpyxl.Workbook
        """
        work_book = openpyxl.Workbook()
        work_sheet = work_book.active
        work_sheet.title = self._get_sheet_name()
        row = 0

        for col, field in enumerate(self._get_header()):
            # проход header
            cell = work_sheet.cell(row=row + 1, column=col + 1)
            cell.value = field
            cell.font = Font(bold=True)

        for row_data in self._get_data():
            # проход данных
            row += 1
            for col, field in enumerate(row_data):
                cell = work_sheet.cell(row=row + 1, column=col + 1)
                cell.value = self._get_field_value(field)

        return work_book


class XLSXFromQueryset(XLSXSimple):
    """ Генерирует xlsx на основе django QuerySet
    """
    def __init__(self, fields_data=None):
        """ Инициализация из queryset """
        assert isinstance(fields_data, QuerySet), 'fields_data is not queryset'
        self.model_fields = [
            field for field in fields_data.model._meta.concrete_fields
        ]
        super(XLSXFromQueryset, self).__init__(fields_data=fields_data)

    def _get_sheet_name(self):
        return self.fields_data.model._meta.verbose_name

    def _get_header(self):
        return [field.verbose_name for field in self.model_fields]

    def _get_data(self):
        res = []
        field_names = [field.name for field in self.model_fields]
        for obj in self.fields_data:
            res.append([getattr(obj, attr) for attr in field_names])

        return res

    @staticmethod
    def _get_field_value(field):
        if isinstance(field, (UUID, Model)):
            return str(field)
        if isinstance(field, (dict, list)):
            return json.dumps(field)

        return field